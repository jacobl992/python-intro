import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions2.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
    #can also use cell = sheet.cell(1, 1)
print(cell.value)
print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 5)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=5,
          max_col=5)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')

#this code creates a new set of values in the table and then makes a barchart from those values
#this is done by using the third-party package openpyxl, installed via terminal with command 'pip3 install openpyxl'
#read more of the documentation to see how to do colour code the bars in the chart, and other things like labels etc
#you could use this code to iterate over many xl files and make instant changes